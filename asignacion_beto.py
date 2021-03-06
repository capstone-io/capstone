import xlrd

from gurobipy import *
import networkx as nx #Para trabajar con grafos
import matplotlib.pyplot as plt #Para visualizar los grafos

from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font,Alignment
from datetime import date, timedelta
import os


# Give the location of the file
loc = ("Datos con asignación.xlsx")

# To open Workbook
wb = xlrd.open_workbook(loc)


sheet_obras = wb.sheet_by_index(0)
sheet_plantas = wb.sheet_by_index(1)
sheet_calculos_ruteo = wb.sheet_by_index(5)

dias = [i+1 for i in range(7)]
obras = [i+1 for i in range(225)]
# Obras demanda separa las obras con demanda mayor a 120 en un dia en dos obras, 
# porque la tienen que despachar dos camiones
obras_demanda = []
for i in range(225):
    mayor_120 = 0
    for dia in dias:
        # Se agregan obras con demanda que pueden caber en un camión grande
        if int(sheet_obras.cell_value(i+4, dia+4)) > 120:
            mayor_120 = 1
    if mayor_120 == 0:
        obras_demanda.append(i+1)
    else:
        obras_demanda.append(i+1)
        obras_demanda.append(i+1.1)

plantas = [i+1 for i in range(4)]
# Cree conjunto dias_inventa para mantener el inventario del 8vo día positivo y que se cumpla la restricción el domingo
dias_inventa = [i+1 for i in range(8)]
turnos = [i+1 for i in range(3)]

# Posiciones de las obras
posiciones_obras = {}
# Producción máxima de cada planta
produccion_max = {}
# Producción máxima de cada planta
demanda_diaria = {}
# Distancia entre obras y plantas
distancia_obra_planta = {}
# Stock en día 1
stock_inicial = {}
# Costos de producción por planta
costos_planta = {}
# Costos de almacenamiento por planta
almacenamiento_planta = {}
# Tiempo maximo de uso de camión en despacho en dia t de obra x
# Se considerada viaje de ida y vuelta
tiempo_uso_camion = {}
# Para cada obra se tiene el tiempo que demora a la planta más lejana
tiempo_viaje_obra_planta = {}
# Duracion de descarga de las obras
duracion_descarga_obras = {}
# Para un dia d, obras que se demoran menos de t horas en ser despachadas
tiempo_despachos = {}
# Turnos de disponibilidad para almacenamiento de las distintas obras
turnos_ab = {}
for obra in obras_demanda:
    for turno in turnos:
        turnos_ab[(obra, turno)] = int(sheet_obras.cell_value(int(obra)+3,turno+11))
        
posiciones_plantas = {}
for planta in plantas:
    posiciones_plantas[planta] = (int(sheet_plantas.cell_value(planta+2, 2)), int(sheet_plantas.cell_value(planta+2, 3)))

for planta in plantas:
    produccion_max[planta] = int(sheet_plantas.cell_value(planta+2, 7))
for planta in plantas:
    stock_inicial[planta] = int(sheet_plantas.cell_value(planta+2, 8))
for planta in plantas:
    costos_planta[planta] = float(sheet_plantas.cell_value(planta+2, 5))
for planta in plantas:
    almacenamiento_planta[planta] = float(sheet_plantas.cell_value(planta+2, 6))

for obra in obras:
    for dia in dias:
        # Se agregan obras con demanda que pueden caber en un camión grande
        if int(sheet_obras.cell_value(obra+3, dia+4)) <= 120:
            demanda_diaria[(obra, dia)]= int(sheet_obras.cell_value(obra+3, dia+4))
            demanda_diaria[(obra+0.1, dia)] = 0
        else:
            if 120 < int(sheet_obras.cell_value(obra+3, dia+4)) <= 130:
                # se usa un camion mediano al maximo y uno chico con 30-40
                demanda_diaria[(obra, dia)]= 90
                demanda_diaria[(obra+0.1, dia)]= int(sheet_obras.cell_value(obra+3, dia+4))- 90
            elif 130 < int(sheet_obras.cell_value(obra+3, dia+4)) <= 140:
                # se usa un camion mediano al maximo y uno chico con 40-50
                demanda_diaria[(obra, dia)]= 90
                demanda_diaria[(obra+0.1, dia)]= int(sheet_obras.cell_value(obra+3, dia+4))- 90
            elif 140 < int(sheet_obras.cell_value(obra+3, dia+4)) <= 150:
                # se usa un camion mediano al maximo y uno chico con 50-60
                demanda_diaria[(obra, dia)]= 90
                demanda_diaria[(obra+0.1, dia)]= int(sheet_obras.cell_value(obra+3, dia+4))- 90
            elif 150 < int(sheet_obras.cell_value(obra+3, dia+4)) <= 160:
                # se usan dos camiones medianos, uno en su maxima capacidad y el otro con 60-70
                demanda_diaria[(obra, dia)]= 90
                demanda_diaria[(obra+0.1, dia)]= int(sheet_obras.cell_value(obra+3, dia+4))- 90
            elif 160 < int(sheet_obras.cell_value(obra+3, dia+4)) <= 170:
                # se usan dos camiones medianos, uno en su maxima capacidad y el otro con 70-80
                demanda_diaria[(obra, dia)]= 90
                demanda_diaria[(obra+0.1, dia)]= int(sheet_obras.cell_value(obra+3, dia+4))- 90
            elif 170 < int(sheet_obras.cell_value(obra+3, dia+4)) <= 180:
                # se usa un camion grande en su maxima capacidad y uno chico en 50-60
                demanda_diaria[(obra, dia)]= 120
                demanda_diaria[(obra+0.1, dia)]= int(sheet_obras.cell_value(obra+3, dia+4))- 120
            elif 180 < int(sheet_obras.cell_value(obra+3, dia+4)) <= 190:
                # se usa un camion grande en su maxima capacidad y uno mediano en 60-70
                demanda_diaria[(obra, dia)]= 120
                demanda_diaria[(obra+0.1, dia)]= int(sheet_obras.cell_value(obra+3, dia+4))- 120
            elif 190 < int(sheet_obras.cell_value(obra+3, dia+4)) <= 200:
                # se usa un camion grande en su maxima capacidad y uno mediano en 70-80
                demanda_diaria[(obra, dia)]= 120
                demanda_diaria[(obra+0.1, dia)]= int(sheet_obras.cell_value(obra+3, dia+4))- 120
            elif 200 < int(sheet_obras.cell_value(obra+3, dia+4)) <= 210:
                # se usa un camion grande en su maxima capacidad y uno mediano en 80-90
                demanda_diaria[(obra, dia)]= 120
                demanda_diaria[(obra+0.1, dia)]= int(sheet_obras.cell_value(obra+3, dia+4))- 120
        #print(demanda_diaria)


for obra in obras_demanda:
    for planta in plantas:
        distancia_obra_planta[(obra, planta)] = int(sheet_obras.cell_value(int(obra)+3, planta+16))
        
for obra in obras_demanda:
    posiciones_obras[obra] = (int(sheet_obras.cell_value(int(obra)+3, 2)), int(sheet_obras.cell_value(int(obra)+3, 3)))

for obra in obras_demanda:
    # El tiempo se divide en 10 para que quede el tiempo de descarga por tonelada de hormigón
    duracion_descarga_obras[obra] = (float(sheet_obras.cell_value(int(obra)+3, 4))/10)

# se usa para restriccion de tiempo de camiones        

for obra in obras_demanda:
    for planta in plantas:
        tiempo_viaje_obra_planta[planta, obra] = float(sheet_calculos_ruteo.cell_value(int(obra)+2, int(planta)+7))
#print(tiempo_viaje_obra_planta)
#print(demanda_diaria)
# se usa para restriccion de tiempo de camiones
for obra in obras_demanda:
    for dia in dias:
        for planta in plantas:
            tiempo_despachos[planta, obra, dia] = tiempo_viaje_obra_planta[planta, obra] + demanda_diaria[obra, dia]*duracion_descarga_obras[obra]
           # print('tiempo, obra, dia, demanda, tiempo viaje, tiempo descarga', tiempo_despacho, obra, dia, demanda_diaria[obra, dia], tiempo_maximo_viaje_obra[obra], duracion_descarga_obras[obra])

#print(tiempo_uso_camion)
#print(float(sheet_calculos_ruteo.cell_value(3, 20)))


m = Model('Asignacion de obras a plantas')

# Variable de asignación. 1 si se le asigna la obra a la planta el un día específico. 0 en otro caso
x = {}

for planta in plantas:
        for obra in obras_demanda:
            for dia in dias:
                    x[(planta, dia, obra)] = m.addVar(vtype=GRB.BINARY, name='x_{}_{}_{}'.format(
                                                                             planta, dia, obra))
m.update()

A = {}

for planta in plantas:
        for obra in obras_demanda:
            for dia in range(1,7):
                    A[(planta, dia, obra)] = m.addVar(vtype=GRB.BINARY, name='A_{}_{}_{}'.format(
                                                                             planta, dia, obra))
m.update()


# Variable de Inventario. Cantidad de inventario disponible al principio del día en una planta.
# Dura hasta el día 8 para asegurar que se cumpla la demanda
Inventario = {}

for planta in plantas:
    for dia in dias_inventa:
        Inventario[planta, dia] = m.addVar(vtype=GRB.CONTINUOUS, name='Inventario_{}_{}'.format(
                                                                             planta, dia))
m.update()

# Variable de producción. Cantidad producida en una planta en un día
Produc = {}

for planta in plantas:
    for dia in range(0, 8):
        Produc[planta, dia] = m.addVar(vtype=GRB.CONTINUOUS, name='Produc_{}_{}'.format(
                                                                             planta, dia))
m.update()



planta_0_1 = 0
obras_asig_0_1 = []
planta_0_2 = 0
obras_asig_0_2 = []
planta_0_3 = 0
obras_asig_0_3 = []
planta_0_4 = 0
obras_asig_0_4 = []
planta_0_5 = 0
obras_asig_0_5 = []
planta_0_6 = 0
obras_asig_0_6 = []
planta_0_7 = 0
obras_asig_0_7 = []

# 1. Si una obra tiene demanda mayor a 0 en un día específico entonces tiene que tener una planta asignada
for obra in obras_demanda:
    for dia in range(2, 8):
        if demanda_diaria[(obra, dia)] > 0:
            m.addConstr(quicksum(x[(planta, dia, obra)] for planta in plantas) +
                        quicksum(A[(planta, dia-1, obra)] for planta in plantas) == 1)
m.update()

for obra in obras_demanda:
    if demanda_diaria[(obra, 1)] > 0:
        m.addConstr(quicksum(x[(planta, 1, obra)] for planta in plantas) == 1)
m.update()
# 2. Si una obra tiene demanda igual a 0 en un día específico entonces no tiene ninguna planta asignada.
# Además se agregan a la lista de plantas sin asignación para ese día específico.
for obra in obras_demanda:
    for dia in dias:
        if demanda_diaria[(obra, dia)] == 0:
            for planta in plantas:
                m.addConstr(x[(planta, dia, obra)] == 0)
            if dia == 1:
                planta_0_1 += 1
                obras_asig_0_1.append(obra)
            elif dia == 2:
                planta_0_2 += 1
                obras_asig_0_2.append(obra)
            elif dia == 3:
                planta_0_3 += 1
                obras_asig_0_3.append(obra)
            elif dia == 4:
                planta_0_4 += 1
                obras_asig_0_4.append(obra)
            elif dia == 5:
                planta_0_5 += 1
                obras_asig_0_5.append(obra)
            elif dia == 6:
                planta_0_6 += 1
                obras_asig_0_6.append(obra)
            elif dia == 7:
                planta_0_7 += 1
                obras_asig_0_7.append(obra)
m.update()

for obra in obras_demanda:
    for dia in range(1, 7):
        if demanda_diaria[(obra, dia)] == 0:
            for planta in plantas:
                m.addConstr(A[(planta, dia, obra)] == 0)
m.update()
# 3. El inventario para todos los día incluyendo el primero de la semana siguiente es mayor a 0
for planta in plantas:
    for dia in dias_inventa:
        m.addConstr(Inventario[planta, dia] >= 0)
m.update()

# 4. El inventario para el primer día es igual al inventario inicial
for planta in plantas:
    m.addConstr(Inventario[planta, 1] == stock_inicial[planta])
m.update()

# 5. El inventario al inicio del día, más lo que se produce en ese día, menos la demanda de la obras
# asignadas ese dia es igual al inventario del día siguiente
for planta in plantas:
    for dia in range(1, 7):
        m.addConstr(Inventario[planta, dia] + Produc[planta, dia] -
                    (quicksum(x[(planta, dia, obra)]*demanda_diaria[obra, dia] for obra in obras_demanda)) -
                    (quicksum(A[(planta, dia, obra)]*demanda_diaria[obra, dia+1] for obra in obras_demanda))
                     == Inventario[planta, dia+1])
m.update()

for planta in plantas:
    m.addConstr(Inventario[planta, 7] + Produc[planta, 7] -
                (quicksum(x[(planta, 7, obra)]*demanda_diaria[obra, 7]
                          for obra in obras_demanda)) == Inventario[planta, 8])
m.update()


# 6. Suplir la demanda de cada día
for planta in plantas:
    m.addConstr(Inventario[planta, 7] >=
                (quicksum(x[(planta, 7, obra)]*demanda_diaria[obra, 7]
                for obra in obras_demanda)))
m.update()

for planta in plantas:
    for dia in range(1, 7):
        m.addConstr(Inventario[planta, dia] >=
                    (quicksum(x[(planta, dia, obra)]*demanda_diaria[obra, dia] +
                              A[(planta, dia, obra)]*demanda_diaria[obra, dia+1]
                 for obra in obras_demanda)))
m.update()

# 7. Producción máxima por planta
for planta in plantas:
    for dia in dias:
        m.addConstr(Produc[planta, dia] <= produccion_max[planta])
m.update()


obj = quicksum((Produc[planta, dia]*costos_planta[planta])
               for planta in plantas for dia in dias)\
      + quicksum((Inventario[planta, dia]*almacenamiento_planta[planta])
                 for dia in range(2,8) for planta in plantas) \
      + quicksum((x[(planta, dia, obra)]*distancia_obra_planta[(obra, planta)]*0.0021)
                 for obra in obras_demanda for planta in plantas for dia in dias) \
      + quicksum((A[(planta, dia, obra)]*distancia_obra_planta[(obra, planta)]*0.0021)
                 for obra in obras_demanda for planta in plantas for dia in range(1,7))\
      + quicksum((A[(planta, dia, obra)]*demanda_diaria[(obra, dia+1)]*0.017)
                 for planta in plantas for obra in obras for dia in range(1, 7))


m.setObjective(obj,GRB.MINIMIZE)
m.update()
m.optimize()



planta_1_1=0
obras_asig_1_1=[]
planta_2_1=0
obras_asig_2_1=[]
planta_3_1=0
obras_asig_3_1=[]
planta_4_1=0
obras_asig_4_1=[]
planta_1_2=0
obras_asig_1_2=[]
planta_2_2=0
obras_asig_2_2=[]
planta_3_2=0
obras_asig_3_2=[]
planta_4_2=0
obras_asig_4_2=[]
planta_1_3=0
obras_asig_1_3=[]
planta_2_3=0
obras_asig_2_3=[]
planta_3_3=0
obras_asig_3_3=[]
planta_4_3=0
obras_asig_4_3=[]
planta_1_4=0
obras_asig_1_4=[]
planta_2_4=0
obras_asig_2_4=[]
planta_3_4=0
obras_asig_3_4=[]
planta_4_4=0
obras_asig_4_4=[]
planta_1_5=0
obras_asig_1_5=[]
planta_2_5=0
obras_asig_2_5=[]
planta_3_5=0
obras_asig_3_5=[]
planta_4_5=0
obras_asig_4_5=[]
planta_1_6=0
obras_asig_1_6=[]
planta_2_6=0
obras_asig_2_6=[]
planta_3_6=0
obras_asig_3_6=[]
planta_4_6=0
obras_asig_4_6=[]
planta_1_7=0
obras_asig_1_7=[]
planta_2_7=0
obras_asig_2_7=[]
planta_3_7=0
obras_asig_3_7=[]
planta_4_7=0
obras_asig_4_7=[]

obras_asignadas = {}
cantidad_obras_asignadas = {}
obras_adelantadas = {}
cantidad_obras_adelantadas = {}
for planta in plantas:
    for dia in dias:
        obras_asignadas[planta, dia] = []
        cantidad_obras_asignadas[planta, dia] = 0
for planta in plantas:
    for dia in dias:
        obras_adelantadas[planta, dia] = []
        cantidad_obras_adelantadas[planta, dia] = 0

produccion_plantas = []
inventario_plantas = {}
costos_prod = 0
costo_inventario = 0
costos_adelantar = 0
for v in m.getVars():
    if v.varName[0:6] == "Produc":
        produccion_plantas.append((v.varName, v.x))
        costos_prod += (v.x*costos_planta[float(v.varName[7:8])])
    if v.varName[0:10] == "Inventario":
        inventario_plantas[(int(v.varName[11:12]), int(v.varName[13:]))]= v.x
        costo_inventario += (v.x*almacenamiento_planta[float(v.varName[11:12])])
    if v.varName[0:1] == "A" and int(v.varName[4:5])<7:
        costos_adelantar += (v.x*0.017*demanda_diaria[(int(v.varName[6:7]),int(v.varName[4:5])+1)])
    if v.varName[0:5]=="x_1_1":
        if v.x==1:
            cantidad_obras_asignadas[1,1]+=1
            obras_asignadas[1, 1].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_2_1":
        if v.x==1:
            cantidad_obras_asignadas[2,1]+=1
            obras_asignadas[2, 1].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_3_1":
        if v.x==1:
            cantidad_obras_asignadas[3,1]+=1
            obras_asignadas[3, 1].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_4_1":
        if v.x==1:
            cantidad_obras_asignadas[4,1]+=1
            obras_asignadas[4, 1].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_1_2":
        if v.x==1:
            cantidad_obras_asignadas[1,2]+=1
            obras_asignadas[1, 2].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_2_2":
        if v.x==1:
            cantidad_obras_asignadas[2,2]+=1
            obras_asignadas[2, 2].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_3_2":
        if v.x==1:
            cantidad_obras_asignadas[3,2]+=1
            obras_asignadas[3, 2].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_4_2":
        if v.x==1:
            cantidad_obras_asignadas[4,2]+=1
            obras_asignadas[4, 2].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_1_3":
        if v.x==1:
            cantidad_obras_asignadas[1,3]+=1
            obras_asignadas[1, 3].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_2_3":
        if v.x==1:
            cantidad_obras_asignadas[2,3]+=1
            obras_asignadas[2, 3].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_3_3":
        if v.x==1:
            cantidad_obras_asignadas[3,3]+=1
            obras_asignadas[3, 3].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_4_3":
        if v.x==1:
            cantidad_obras_asignadas[4,3]+=1
            obras_asignadas[4, 3].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_1_4":
        if v.x==1:
            cantidad_obras_asignadas[1,4]+=1
            obras_asignadas[1, 4].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_2_4":
        if v.x==1:
            cantidad_obras_asignadas[2,4]+=1
            obras_asignadas[2, 4].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_3_4":
        if v.x==1:
            cantidad_obras_asignadas[3,4]+=1
            obras_asignadas[3, 4].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_4_4":
        if v.x==1:
            cantidad_obras_asignadas[4,4]+=1
            obras_asignadas[4, 4].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_1_5":
        if v.x==1:
            cantidad_obras_asignadas[1,5]+=1
            obras_asignadas[1, 5].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_2_5":
        if v.x==1:
            cantidad_obras_asignadas[2,5]+=1
            obras_asignadas[2, 5].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_3_5":
        if v.x==1:
            cantidad_obras_asignadas[3,5]+=1
            obras_asignadas[3, 5].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_4_5":
        if v.x==1:
            cantidad_obras_asignadas[4,5]+=1
            obras_asignadas[4, 5].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_1_6":
        if v.x==1:
            cantidad_obras_asignadas[1,6]+=1
            obras_asignadas[1, 6].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_2_6":
        if v.x==1:
            cantidad_obras_asignadas[2,6]+=1
            obras_asignadas[2, 6].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_3_6":
        if v.x==1:
            cantidad_obras_asignadas[3,6]+=1
            obras_asignadas[3, 6].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_4_6":
        if v.x==1:
            cantidad_obras_asignadas[4,6]+=1
            obras_asignadas[4, 6].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_1_7":
        if v.x==1:
            cantidad_obras_asignadas[1,7]+=1
            obras_asignadas[1, 7].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_2_7":
        if v.x==1:
            cantidad_obras_asignadas[2,7]+=1
            obras_asignadas[2, 7].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_3_7":
        if v.x==1:
            cantidad_obras_asignadas[3,7]+=1
            obras_asignadas[3, 7].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="x_4_7":
        if v.x==1:
            cantidad_obras_asignadas[4,7]+=1
            obras_asignadas[4, 7].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_1_1":
        if v.x==1:
            cantidad_obras_adelantadas[1,1]+=1
            obras_adelantadas[1, 1].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_2_1":
        if v.x==1:
            cantidad_obras_adelantadas[2,1]+=1
            obras_adelantadas[2, 1].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_3_1":
        if v.x==1:
            cantidad_obras_adelantadas[3,1]+=1
            obras_adelantadas[3, 1].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_4_1":
        if v.x==1:
            cantidad_obras_adelantadas[4,1]+=1
            obras_adelantadas[4, 1].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_1_2":
        if v.x==1:
            cantidad_obras_adelantadas[1,2]+=1
            obras_adelantadas[1, 2].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_2_2":
        if v.x==1:
            cantidad_obras_adelantadas[2,2]+=1
            obras_adelantadas[2, 2].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_3_2":
        if v.x==1:
            cantidad_obras_adelantadas[3,2]+=1
            obras_adelantadas[3, 2].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_4_2":
        if v.x==1:
            cantidad_obras_adelantadas[4,2]+=1
            obras_adelantadas[3, 2].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_1_3":
        if v.x==1:
            cantidad_obras_adelantadas[1,3]+=1
            obras_adelantadas[1, 3].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_2_3":
        if v.x==1:
            cantidad_obras_adelantadas[2,3]+=1
            obras_adelantadas[2, 3].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_3_3":
        if v.x==1:
            cantidad_obras_adelantadas[3,3]+=1
            obras_adelantadas[3, 3].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_4_3":
        if v.x==1:
            cantidad_obras_adelantadas[4,3]+=1
            obras_adelantadas[4, 3].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_1_4":
        if v.x==1:
            cantidad_obras_adelantadas[1,4]+=1
            obras_adelantadas[1, 4].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_2_4":
        if v.x==1:
            cantidad_obras_adelantadas[2,4]+=1
            obras_adelantadas[2, 4].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_3_4":
        if v.x==1:
            cantidad_obras_adelantadas[3,4]+=1
            obras_adelantadas[3, 4].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_4_4":
        if v.x==1:
            cantidad_obras_adelantadas[4,4]+=1
            obras_adelantadas[4, 4].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_1_5":
        if v.x==1:
            cantidad_obras_adelantadas[1,5]+=1
            obras_adelantadas[1, 5].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_2_5":
        if v.x==1:
            cantidad_obras_adelantadas[2,5]+=1
            obras_adelantadas[2, 5].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_3_5":
        if v.x==1:
            cantidad_obras_adelantadas[3,5]+=1
            obras_adelantadas[3, 5].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_4_5":
        if v.x==1:
            cantidad_obras_adelantadas[4,5]+=1
            obras_adelantadas[4, 5].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_1_6":
        if v.x==1:
            cantidad_obras_adelantadas[1,6]+=1
            obras_adelantadas[1, 6].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_2_6":
        if v.x==1:
            cantidad_obras_adelantadas[2,6]+=1
            obras_adelantadas[2, 6].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_3_6":
        if v.x==1:
            cantidad_obras_adelantadas[3,6]+=1
            obras_adelantadas[3, 6].append(float(v.varName[6:len(v.varName)]))
    elif v.varName[0:5]=="A_4_6":
        if v.x==1:
            cantidad_obras_adelantadas[4,6]+=1
            obras_adelantadas[4, 6].append(float(v.varName[6:len(v.varName)]))

"""
print("")    
print("Planta 1 abastece ",planta_1_1," obras en el día 1")
print("Obras asignadas a P1 en el día 1",obras_asig_1_1)
print("")
print("Planta 2 abastece ",planta_2_1," obras en el día 1")
print("Obras asignadas a P2 en el día 1",obras_asig_2_1)
print("")
print("Planta 3 abastece ",planta_3_1," obras en el día 1")
print("Obras asignadas a P3 en el día 1",obras_asig_3_1)
print("")
print("Planta 4 abastece ",planta_4_1," obras en el día 1")
print("Obras asignadas a P4 en el día 1",obras_asig_4_1)
print("")
print(planta_0_1," obras no tienen demanda en el día 1")
print("Obras sin demanda en el día 1", obras_asig_0_1)
print("")
print("")
print("Planta 1 abastece ",planta_1_2," obras en el día 2")
print("Obras asignadas a P1 en el día 2",obras_asig_1_2)
print("")
print("Planta 2 abastece ",planta_2_2," obras en el día 2")
print("Obras asignadas a P2 en el día 2",obras_asig_2_2)
print("")
print("Planta 3 abastece ",planta_3_2," obras en el día 2")
print("Obras asignadas a P3 en el día 2",obras_asig_3_2)
print("")
print("Planta 4 abastece ",planta_4_2," obras en el día 2")
print("Obras asignadas a P4 en el día 2",obras_asig_4_2)
print("")
print(planta_0_2, " obras no tienen demanda en el día 2")
print("Obras sin demanda en el día 2", obras_asig_0_2)
print("")
print("")
print("Planta 1 abastece ",planta_1_3," obras en el día 3")
print("Obras asignadas a P1 en el día 3",obras_asig_1_3)
print("")
print("Planta 2 abastece ",planta_2_3," obras en el día 3")
print("Obras asignadas a P2 en el día 3",obras_asig_2_3)
print("")
print("Planta 3 abastece ",planta_3_3," obras en el día 3")
print("Obras asignadas a P3 en el día 3",obras_asig_3_3)
print("")
print("Planta 4 abastece ",planta_4_3," obras en el día 3")
print("Obras asignadas a P4 en el día 3",obras_asig_4_3)
print("")
print(planta_0_3," obras no tienen demanda en el día 3")
print("Obras sin demanda en el día 3", obras_asig_0_3)
print("")
print("")
print("Planta 1 abastece ",planta_1_4," obras en el día 4")
print("Obras asignadas a P1 en el día 4",obras_asig_1_4)
print("")
print("Planta 2 abastece ",planta_2_4," obras en el día 4")
print("Obras asignadas a P2 en el día 4",obras_asig_2_4)
print("")
print("Planta 3 abastece ",planta_3_4," obras en el día 4")
print("Obras asignadas a P3 en el día 4",obras_asig_3_4)
print("")
print("Planta 4 abastece ",planta_4_4," obras en el día 4")
print("Obras asignadas a P4 en el día 4",obras_asig_4_4)
print("")
print(planta_0_4," obras no tienen demanda en el día 4")
print("Obras sin demanda en el día 4", obras_asig_0_4)
print("")
print("")
print("Planta 1 abastece ",planta_1_5," obras en el día 5")
print("Obras asignadas a P1 en el día 5",obras_asig_1_5)
print("")
print("Planta 2 abastece ",planta_2_5," obras en el día 5")
print("Obras asignadas a P2 en el día 5",obras_asig_2_5)
print("")
print("Planta 3 abastece ",planta_3_5," obras en el día 5")
print("Obras asignadas a P3 en el día 5",obras_asig_3_5)
print("")
print("Planta 4 abastece ",planta_4_5," obras en el día 5")
print("Obras asignadas a P4 en el día 5",obras_asig_4_5)
print("")
print(planta_0_5," obras no tienen demanda en el día 5")
print("Obras sin demanda en el día 5", obras_asig_0_5)
print("")
print("")
print("Planta 1 abastece ",planta_1_6," obras en el día 6")
print("Obras asignadas a P1 en el día 6",obras_asig_1_6)
print("")
print("Planta 2 abastece ",planta_2_6," obras en el día 6")
print("Obras asignadas a P2 en el día 6",obras_asig_2_6)
print("")
print("Planta 3 abastece ",planta_3_6," obras en el día 6")
print("Obras asignadas a P3 en el día 6",obras_asig_3_6)
print("")
print("Planta 4 abastece ",planta_4_6," obras en el día 6")
print("Obras asignadas a P4 en el día 6",obras_asig_4_6)
print("")
print(planta_0_6," obras no tienen demanda en el día 6")
print("Obras sin demanda en el día 6", obras_asig_0_6)
print("")
print("")
print("Planta 1 abastece ",planta_1_7," obras en el día 7")
print("Obras asignadas a P1 en el día 7",obras_asig_1_6)
print("")
print("Planta 2 abastece ",planta_2_7," obras en el día 7")
print("Obras asignadas a P2 en el día 7",obras_asig_2_6)
print("")
print("Planta 3 abastece ",planta_3_7," obras en el día 7")
print("Obras asignadas a P3 en el día 7",obras_asig_3_6)
print("")
print("Planta 4 abastece ",planta_4_7," obras en el día 7")
print("Obras asignadas a P4 en el día 7",obras_asig_4_6)
print("")
print(planta_0_7," obras no tienen demanda en el día 7")
print("Obras sin demanda en el día 7", obras_asig_0_7)
print("")
print("Producción de plantas", produccion_plantas)
print("Inventario plantas", inventario_plantas)
"""

#lib=load_workbook(filename="Datos.xlsx")                                                    #
#del lib["Asignación"]
#ws = lib.create_sheet("Asignación",index=3)#

#thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    # top=Side(style='thin'), bottom=Side(style='thin'))

#cell = ws.cell(row=1, column=1, value="Obras")
#cell.font = Font(bold=True, )
#cell.border = thin_border
#cell = ws.cell(row=1, column=2, value="Planta 1")
#cell.font = Font(bold=True, )
#cell.border = thin_border
#cell = ws.cell(row=1, column=3, value="Planta 2")
#    cell.font = Font(bold=True, )
 #   cell.border = thin_border
  #  cell = ws.cell(row=1, column=4, value="Planta 3")
   # cell.font = Font(bold=True, )
    #cell.border = thin_border
#    cell = ws.cell(row=1, column=5, value="Planta 4")
 #   cell.font = Font(bold=True, )
  #  cell.border = thin_border

   # for i in range(225):
    #    if (i+1) in obras_asig_1:
     #       cell = ws.cell(row=2+i, column=2, value=1)
      #      cell = ws.cell(row=2+i, column=3, value=0)
       #     cell = ws.cell(row=2+i, column=4, value=0)
        #    cell = ws.cell(row=2+i, column=5, value=0)
        #elif (i+1) in obras_asig_2:
         #   cell = ws.cell(row=2+i, column=2, value=0)
          #  cell = ws.cell(row=2+i, column=3, value=1)
           # cell = ws.cell(row=2+i, column=4, value=0)
            #cell = ws.cell(row=2+i, column=5, value=0)
 #       elif (i+1) in obras_asig_3:
  #          cell = ws.cell(row=2+i, column=2, value=0)
   #         cell = ws.cell(row=2+i, column=3, value=0)
    #        cell = ws.cell(row=2+i, column=4, value=1)
     #       cell = ws.cell(row=2+i, column=5, value=0)
      #  else:
       #     cell = ws.cell(row=2+i, column=2, value=0)
        #    cell = ws.cell(row=2+i, column=3, value=0)
         #   cell = ws.cell(row=2+i, column=4, value=0)
          #  cell = ws.cell(row=2+i, column=5, value=1)

#    for i in range(225):
 #       cell = ws.cell(row=2+i, column=1, value="Obra "+str(i+1))
  #      cell.border = thin_border

   # try:
    #    lib.save("Datos con asignación.xlsx")
 #   except Excpetion as e:
  #      print(e)
   #     print("Ha ocurrido un problema al intentar guardar el archivo")
"""   
print(obras_asignadas)
print(cantidad_obras_asignadas)
print(obras_adelantadas)
print(cantidad_obras_adelantadas)
print(costos_prod)
print(costo_inventario)
print(costos_adelantar)
print('Costo total: %g' % m.objVal)
print("Producción de plantas", produccion_plantas)
print("Inventario plantas", inventario_plantas)
total = 0
for i in produccion_plantas:
    total += i[1]
print(total)
totalinventario = 0
for i in inventario_plantas.values():
    totalinventario += i
print(totalinventario)
stock = 0
for i in stock_inicial.values():
    stock += i
print(stock)
print(totalinventario-total-stock)
demanda=0
for k, a in obras_adelantadas.items():
    for i in a:
        demanda += demanda_diaria[i, (float(k[1]))]
for k, a in obras_asignadas.items():
    for i in a:
        demanda += demanda_diaria[(i, float(k[1]))]
pedidos1 = 0
for i in demanda_diaria.values():
    if i>0:
        pedidos1+= 1
print(pedidos1)
pedidos = 0
for i in cantidad_obras_adelantadas.values():
    pedidos += i
for i in cantidad_obras_asignadas.values():
    pedidos += i
print(pedidos)
"""